#####################################################
# 	Script for: Converting data in given format		#			 
# 	Output: Data value in given format				#
# 	Input: XL file and format required   	 		#
#####################################################


import binascii
import sys
import openpyxl
from openpyxl.styles import PatternFill

WB = openpyxl.load_workbook('partref.xlsx')
sheet = WB["partRef"]

for r in range(2,12):
    for c in range(1,5):
        each = sheet.cell(row=r, column = c)
        regionC = each.value
        x = binascii.hexlify(regionC)
        tarEach = sheet.cell(row=r+17, column = c)
        if(c == 1):
            tarEach.value = regionC
        if(c == 2):
            tarEach.value = '62F012' + x
        if(c == 3):
            tarEach.value = '62F013' + x
        if(c == 4):
            tarEach.value = '62F188' + x
            F188 = sheet.cell(row=r+17, column = c+1)
            F188.value = '62F1A2' + x
          
WB.save('partref.xlsx')

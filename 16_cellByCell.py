#####################################################
# 	Script for: Verify and alrert for any variation	#			 
#				in given two Excel files     		#	
# 	Output: Highlight cell with different values	#
# 	Input files: XL files							#
#####################################################

import os
import sys
import openpyxl
from openpyxl.styles import PatternFill

masterWB = openpyxl.load_workbook('00r-AIVC-SOP2_sw09.09.08_master.xlsm',read_only=False,keep_vba=True)
Msheet = masterWB["5.0 Data Cross Reference"]
    
userWB = openpyxl.load_workbook('00r-AIVC-SOP2_sw09.09.08.xlsm',read_only=False,keep_vba=True)
Usheet = userWB["5.0 Data Cross Reference"]

MasterRow = 4
UserRow = 3
    
for i in range(1,120):
    MasterCell = Msheet.cell(row=MasterRow, column = i)
    UserCell = Usheet.cell(row=UserRow, column = i) 
    MasterValue = MasterCell.value
    UserValue = UserCell.value
    col = openpyxl.utils.get_column_letter(i)
    if(MasterValue != UserValue):
        UserCell.fill = PatternFill(start_color='ffff33', end_color='ffff33',fill_type = 'solid')
        e = Msheet.cell(row=1, column = i)
        regionE = e.value
        print("Found a mismatch in cell: " +'"'+ str(col) + str(UserRow)+'"'+ " for " +'"'+ str(regionE)+ '"')
        print("Master Value: " + str(MasterValue))
        print("User Value: " + str(UserValue))
        print("\n")
    
userWB.save('00r-AIVC-SOP2_sw09.09.08.xlsm')
masterWB.save('00r-AIVC-SOP2_sw09.09.08_master.xlsm')
  
'''
1. Check coloumn one by one 
2. if any change in the cell - fill the cell "yellow"
3. Then move to the next row.
4. complete the full cell

ws.cell(row = 4, column = 2)

'''
